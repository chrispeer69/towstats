"""Deterministic generator for a realistic Towbook "Digital Requests" export.

This module is the reason the whole pipeline is testable without portal access.
It writes an XLSX that looks like the real Export -> Excel output -- the same
header names, the same cell types, the same messiness -- so ingestion,
classification, metrics, alerting and the dashboard can all be exercised offline.

Two entry points matter:

``generate_fixture_xlsx(path, days=14, seed=42)``
    The big realistic dataset. Deterministic for a given ``seed`` and
    ``end_date``, so tests can assert exact numbers.

``write_rows_xlsx(path, rows)``
    The low-level writer. Hand-authored rows in, XLSX out. Used by the tests
    that need a handful of rows with numbers that were worked out by hand.

Header names are read from ``config/schema.yaml`` (first candidate per canonical
field) rather than hardcoded, so that reconciling the schema against a real
export automatically reconciles the fixture too.

What the generated dataset deliberately contains
------------------------------------------------
* several clients with clearly different acceptance rates;
* one client (:data:`DEGRADING_CLIENT`) whose rate collapses in the most recent
  seven days -- enough offers in the final 24h to trip ``client_acceptance_drop``
  and a big enough swing to show up as a weekly outlier;
* tows that were denied or expired, which is what ``missed_tow`` fires on;
* light service that was *accepted* against the acceptance policy;
* at least :data:`MIN_UNCLASSIFIED` rows across at least two distinct service
  type strings that match no rule at all, so ``unclassified`` is exercised;
* free-text denial reasons, unbucketed, exactly as the portal supplies them;
* a realistic hour-of-day curve with a pronounced overnight dip;
* mixed cell types on purpose: ``Date and Time`` is a real datetime cell while
  ``Responded`` is a formatted string, and some ``Amount`` cells arrive as
  ``"$1,234.50"`` -- both paths through the ingester get walked.

Run it directly to produce a file::

    python tests/fixture_generator.py --out raw/fixture.xlsx --days 14 --seed 42
"""

from __future__ import annotations

import argparse
import random
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook

__all__ = [
    "generate_fixture_xlsx",
    "generate_rows",
    "write_rows_xlsx",
    "summarize_rows",
    "header_for",
    "headers_for_fields",
    "load_schema",
    "load_rules",
    "service_class_for",
    "canonical_status",
    "CLIENTS",
    "DEGRADING_CLIENT",
    "UNCLASSIFIED_SERVICE_TYPES",
    "FIELD_ORDER",
    "DEFAULT_HEADERS",
    "RECENT_WINDOW_DAYS",
    "MIN_UNCLASSIFIED",
    "MIN_MISSED_TOWS",
    "MIN_LIGHT_SERVICE_ACCEPTED",
]

# --------------------------------------------------------------------------
# Repo location. This file lives at <repo>/tests/fixture_generator.py.
# --------------------------------------------------------------------------

_REPO_ROOT = Path(__file__).resolve().parents[1]
_FALLBACK_CONFIG_DIR = _REPO_ROOT / "config"


# --------------------------------------------------------------------------
# Config access -- prefer the real loader, fall back to reading the YAML
# directly so the generator still works standalone.
# --------------------------------------------------------------------------


def _read_yaml(path: Path) -> dict[str, Any]:
    import yaml

    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
    except OSError:
        return {}
    return data if isinstance(data, dict) else {}


def load_schema() -> dict[str, Any]:
    """Return config/schema.yaml, via the hot-reloading loader when available."""
    try:
        from towbook_agent.core.config_loader import get_schema

        schema = get_schema()
        if schema:
            return schema
    except Exception:  # pragma: no cover - standalone use without the package
        pass
    return _read_yaml(_FALLBACK_CONFIG_DIR / "schema.yaml")


def load_rules() -> dict[str, Any]:
    """Return config/rules.yaml, via the hot-reloading loader when available."""
    try:
        from towbook_agent.core.config_loader import get_rules

        rules = get_rules()
        if rules:
            return rules
    except Exception:  # pragma: no cover - standalone use without the package
        pass
    return _read_yaml(_FALLBACK_CONFIG_DIR / "rules.yaml")


# --------------------------------------------------------------------------
# Header names
# --------------------------------------------------------------------------

#: Column order mirrors the documented Digital Requests grid, then the extra
#: columns the canonical record needs. Entries that are plain strings are
#: literal headers with no canonical field behind them.
FIELD_ORDER: tuple[str, ...] = (
    "offered_at",
    "request_id",
    "client_name",
    "Call Number",
    "PO#",
    "service_type_raw",
    "Vehicle",
    "Request Expiration Time",
    "status",
    "responded_at",
    "denial_reason",
    "pickup_location",
    "dropoff_location",
    "truck_assigned",
    "driver_assigned",
    "amount",
)

#: Used only when config/schema.yaml cannot be read.
DEFAULT_HEADERS: dict[str, str] = {
    "request_id": "ID",
    "client_name": "Provider/Contractor ID",
    "offered_at": "Date and Time",
    "responded_at": "Responded",
    "status": "Status",
    "denial_reason": "Denial Reason",
    "service_type_raw": "Service Needed",
    "pickup_location": "Pickup Location",
    "dropoff_location": "Dropoff Location",
    "truck_assigned": "Truck",
    "driver_assigned": "Driver",
    "amount": "Amount",
}


def header_for(field_name: str, schema: dict[str, Any] | None = None) -> str:
    """Return the export header the ingester expects for a canonical field.

    The first candidate in ``schema.yaml -> columns[field]`` wins, because the
    ingester treats candidate order as priority order. Anything that is not a
    canonical field is returned unchanged and used as a literal header.
    """
    schema = load_schema() if schema is None else schema
    columns = schema.get("columns") or {}
    candidates = columns.get(field_name)
    if isinstance(candidates, str):
        return candidates
    if isinstance(candidates, (list, tuple)) and candidates:
        return str(candidates[0])
    return DEFAULT_HEADERS.get(field_name, field_name)


def headers_for_fields(
    fields: Sequence[str], schema: dict[str, Any] | None = None
) -> list[str]:
    """Map an ordered sequence of canonical fields to export header names."""
    schema = load_schema() if schema is None else schema
    return [header_for(name, schema) for name in fields]


# --------------------------------------------------------------------------
# Test oracles.
#
# These are a deliberately independent reimplementation of the two pieces of
# semantics the fixture depends on. They exist so a test can compute what the
# answer *should* be without asking the code under test. The pipeline must
# never import them.
# --------------------------------------------------------------------------


def service_class_for(service_type_raw: str, rules: dict[str, Any] | None = None) -> str:
    """Independent implementation of the classifier, for building expectations.

    Case-insensitive substring match, file order is evaluation order, first
    match wins, ``_default`` when nothing matches.
    """
    rules = load_rules() if rules is None else rules
    classes = rules.get("service_classes") or {}
    haystack = (service_type_raw or "").casefold()
    for name, spec in classes.items():
        if name.startswith("_"):
            continue
        if not isinstance(spec, dict):
            continue
        for needle in spec.get("match_any") or []:
            if str(needle).casefold() in haystack:
                return name
    default = classes.get("_default", "unclassified")
    return str(default) if default else "unclassified"


def canonical_status(source_status: str, schema: dict[str, Any] | None = None) -> str | None:
    """Independent implementation of the status vocabulary lookup."""
    schema = load_schema() if schema is None else schema
    vocabulary = schema.get("status_vocabulary") or {}
    key = " ".join(str(source_status or "").split()).lower()
    value = vocabulary.get(key)
    return str(value) if value else None


# --------------------------------------------------------------------------
# The shape of the generated dataset
# --------------------------------------------------------------------------

#: How many trailing days count as "recent" for the degradation story.
RECENT_WINDOW_DAYS: int = 7

#: Floors the generator guarantees regardless of how the RNG falls, so tests
#: can assert on them without being flaky.
MIN_UNCLASSIFIED: int = 12
MIN_MISSED_TOWS: int = 10
MIN_LIGHT_SERVICE_ACCEPTED: int = 8
#: Offers made near the end of the export window that have no outcome yet.
MIN_PENDING: int = 3

#: Name of the client whose acceptance rate collapses in the recent window.
DEGRADING_CLIENT: str = "Roadside Rescue Network"


@dataclass(frozen=True)
class ClientProfile:
    """One motor club / dealer sending us work."""

    name: str
    base_rate: float
    volume: tuple[int, int]
    #: acceptance rate inside the recent window; ``None`` means unchanged
    recent_rate: float | None = None
    #: daily volume inside the recent window; ``None`` means unchanged
    recent_volume: tuple[int, int] | None = None

    def rate_on(self, recent: bool) -> float:
        if recent and self.recent_rate is not None:
            return self.recent_rate
        return self.base_rate

    def volume_on(self, recent: bool) -> tuple[int, int]:
        if recent and self.recent_volume is not None:
            return self.recent_volume
        return self.volume


#: Order matters: it is part of what makes the output deterministic.
CLIENTS: tuple[ClientProfile, ...] = (
    ClientProfile("Agero", base_rate=0.90, volume=(14, 20)),
    ClientProfile("Quest Roadside", base_rate=0.71, volume=(8, 13)),
    # The story client: healthy for the first week, falls off a cliff in the
    # most recent seven days, and keeps enough volume in the final 24h to trip
    # client_acceptance_drop (needs >= 10 offers and a rate under 0.60).
    ClientProfile(
        DEGRADING_CLIENT,
        base_rate=0.86,
        volume=(11, 15),
        recent_rate=0.28,
        recent_volume=(13, 18),
    ),
    ClientProfile("Allstate Motor Club", base_rate=0.64, volume=(6, 11)),
    ClientProfile("Copart Transport", base_rate=0.34, volume=(4, 8)),
    ClientProfile("Geico Emergency Road Service", base_rate=0.94, volume=(3, 6)),
)

#: (verbatim service string, weight). The verbatim strings are the point --
#: mixed case, punctuation and trailing detail, because that is what arrives.
SERVICE_TYPES: tuple[tuple[str, int], ...] = (
    # -> tow
    ("Tow", 10),
    ("Light Duty Tow", 9),
    ("Heavy Duty Tow - Accident", 5),
    ("Flatbed Tow", 7),
    ("Accident Recovery", 4),
    ("Wrecker Service", 3),
    ("TOW - PRIVATE PROPERTY IMPOUND", 3),
    # -> winch_out
    ("Winch Out", 5),
    ("Winch-Out / Extraction", 3),
    ("Vehicle Stuck In Ditch", 3),
    ("winchout", 1),
    # -> light_service
    ("Tire Change", 4),
    ("Flat Tire - Spare On Board", 3),
    ("Jump Start", 4),
    ("Battery Boost", 2),
    ("Lockout", 3),
    ("Lock Out Service", 1),
    ("Fuel Delivery", 2),
    ("Out Of Gas", 1),
    # -> unclassified: none of these contain any match_any substring
    ("Motorcycle Transport", 3),
    ("Equipment Haul", 2),
    ("Storage Release", 2),
    ("Mobile Repair", 2),
)

#: Kept as data so a test can assert the fixture really does exercise the
#: mandatory ``unclassified`` default with more than one distinct string.
UNCLASSIFIED_SERVICE_TYPES: tuple[str, ...] = (
    "Motorcycle Transport",
    "Equipment Haul",
    "Storage Release",
    "Mobile Repair",
)

#: Source status strings, per canonical outcome. Every one of them is present
#: in schema.yaml -> status_vocabulary, so the fixture never trips
#: unknown_status_action by accident.
STATUS_STRINGS: dict[str, tuple[str, ...]] = {
    "accepted": ("Accepted", "Completed", "Dispatched", "Tow Complete"),
    "denied": ("Denied", "Declined", "Rejected"),
    "expired": ("Expired", "No Response", "Timed Out"),
    "canceled": ("Canceled", "Canceled By Client", "GOA"),
    "pending": ("Pending", "Open"),
}

#: Free text, unbucketed, exactly as a dispatcher types it.
DENIAL_REASONS: tuple[str, ...] = (
    "No trucks available",
    "All units out on calls",
    "Too far outside our coverage area",
    "Rate too low for a heavy duty run",
    "No driver available - short staffed",
    "Wrong equipment for this vehicle",
    "Already covered by another provider",
    "Cannot handle a class 8 unit",
    "Customer went with another provider",
    "no truck",
    "Declined, light service only - not worth the run",
)

PICKUP_LOCATIONS: tuple[str, ...] = (
    "I-75 NB @ Mile 42, Flint MI",
    "1200 Woodward Ave, Detroit MI",
    "M-59 & Crooks Rd, Rochester Hills MI",
    "4400 Telegraph Rd, Dearborn MI",
    "I-96 WB @ Exit 183, Livonia MI",
    "300 S Main St, Ann Arbor MI",
    "US-23 SB @ Mile 60, Brighton MI",
    "8 Mile & Van Dyke, Warren MI",
)

DROPOFF_LOCATIONS: tuple[str, ...] = (
    "Certified Collision, Troy MI",
    "Bob Maxey Ford, Detroit MI",
    "Owner Residence",
    "City Impound Lot, Pontiac MI",
    "Belle Tire, Sterling Heights MI",
    "Suburban Chrysler, Ann Arbor MI",
)

TRUCKS: tuple[str, ...] = ("Unit 3", "Unit 7", "Unit 12", "T-2 Flatbed", "T-5 Heavy", "Unit 21")
DRIVERS: tuple[str, ...] = (
    "M. Alvarez",
    "D. Kowalski",
    "R. Boone",
    "T. Nguyen",
    "J. Whitfield",
    "C. Ramirez",
)
VEHICLES: tuple[str, ...] = (
    "2019 Ford F-150",
    "2015 Chevrolet Malibu",
    "2021 Ram 2500",
    "2008 Honda Civic",
    "2017 Jeep Grand Cherokee",
    "2013 Freightliner Cascadia",
    "2020 Tesla Model 3",
)

#: Relative offer volume by local hour. Nights are quiet, the morning commute
#: and the afternoon are busy. Index is the hour, 0..23.
HOUR_WEIGHTS: tuple[int, ...] = (
    2, 1, 1, 1, 2, 4,      # 00-05 overnight dip
    7, 12, 14, 13, 12, 11,  # 06-11 morning
    11, 12, 13, 13, 14, 13,  # 12-17 afternoon
    10, 8, 6, 5, 4, 3,      # 18-23 evening tail
)

#: Probability that a light service job is accepted even though the acceptance
#: policy says decline. This is the policy-violation signal the reports surface.
LIGHT_SERVICE_ACCEPT_PROBABILITY: float = 0.18

_DATETIME_OUT_FORMAT = "%m/%d/%Y %I:%M:%S %p"
_EXCEL_DATETIME_FORMAT = "m/d/yyyy h:mm:ss AM/PM"


# --------------------------------------------------------------------------
# Row generation
# --------------------------------------------------------------------------


@dataclass
class _Row:
    """Intermediate row, still in canonical-field terms and local time."""

    request_id: str
    client_name: str
    offered_at: datetime
    responded_at: datetime | None
    status_source: str
    outcome: str
    service_type_raw: str
    service_class: str
    denial_reason: str | None
    pickup_location: str
    dropoff_location: str | None
    truck_assigned: str | None
    driver_assigned: str | None
    amount: float | None
    call_number: str
    po_number: str
    vehicle: str
    expires_at: datetime
    extras: dict[str, Any] = field(default_factory=dict)


def _weighted_choice(rng: random.Random, options: Sequence[tuple[Any, int]]) -> Any:
    values = [value for value, _ in options]
    weights = [weight for _, weight in options]
    return rng.choices(values, weights=weights, k=1)[0]


def _pick_hour(rng: random.Random) -> int:
    return rng.choices(range(24), weights=HOUR_WEIGHTS, k=1)[0]


def _accept_probability(service_class: str, client_rate: float) -> float:
    """How likely this offer is accepted, given the client and the work type."""
    if service_class == "light_service":
        # Policy says decline. Some get taken anyway -- that is the finding.
        return LIGHT_SERVICE_ACCEPT_PROBABILITY
    if service_class == "tow":
        return min(0.97, client_rate * 1.05)
    if service_class == "winch_out":
        return client_rate
    return client_rate * 0.85


def _amount_for(rng: random.Random, service_class: str) -> float:
    ranges = {
        "tow": (95.0, 450.0),
        "winch_out": (150.0, 600.0),
        "light_service": (55.0, 125.0),
    }
    low, high = ranges.get(service_class, (80.0, 300.0))
    return round(rng.uniform(low, high), 2)


def generate_rows(
    days: int = 14,
    seed: int = 42,
    *,
    end_date: date | None = None,
    rules: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Build the fixture dataset as a list of canonical-field row dicts.

    Values are in *local* time (schema.yaml ``source_timezone``), exactly as the
    portal renders them. Deterministic for a given ``seed`` and ``end_date``.
    """
    if days < 1:
        raise ValueError("days must be at least 1")

    rules = load_rules() if rules is None else rules
    rng = random.Random(seed)
    last_day = end_date or date.today()
    first_day = last_day - timedelta(days=days - 1)

    rows: list[_Row] = []
    sequence = 0

    for offset in range(days):
        day = first_day + timedelta(days=offset)
        days_from_end = (last_day - day).days
        recent = days_from_end < RECENT_WINDOW_DAYS
        is_final_day = days_from_end == 0
        # Weekends are quieter for commercial work and busier for consumer work;
        # a single modest factor is enough realism here.
        weekend_factor = 0.8 if day.weekday() >= 5 else 1.0

        for client in CLIENTS:
            low, high = client.volume_on(recent)
            volume = max(1, int(round(rng.randint(low, high) * weekend_factor)))
            client_rate = client.rate_on(recent)

            for _ in range(volume):
                sequence += 1
                hour = _pick_hour(rng)
                offered_at = datetime(
                    day.year,
                    day.month,
                    day.day,
                    hour,
                    rng.randint(0, 59),
                    rng.randint(0, 59),
                )
                service_type_raw = _weighted_choice(rng, SERVICE_TYPES)
                service_class = service_class_for(service_type_raw, rules)

                accepted = rng.random() < _accept_probability(service_class, client_rate)
                if accepted:
                    outcome = "accepted"
                else:
                    outcome = _weighted_choice(
                        rng, (("denied", 6), ("expired", 3), ("canceled", 1))
                    )

                # A handful of the newest offers are still live.
                if is_final_day and hour >= 21 and rng.random() < 0.45:
                    outcome = "pending"

                rows.append(
                    _build_row(
                        rng,
                        sequence=sequence,
                        client=client,
                        offered_at=offered_at,
                        service_type_raw=service_type_raw,
                        service_class=service_class,
                        outcome=outcome,
                    )
                )

    rows.sort(key=lambda row: (row.offered_at, row.request_id))
    _enforce_invariants(rows, rng, rules)
    return [_as_dict(row) for row in rows]


def _build_row(
    rng: random.Random,
    *,
    sequence: int,
    client: ClientProfile,
    offered_at: datetime,
    service_type_raw: str,
    service_class: str,
    outcome: str,
) -> _Row:
    status_source = rng.choice(STATUS_STRINGS[outcome])
    expires_at = offered_at + timedelta(minutes=8)

    if outcome == "pending":
        responded_at: datetime | None = None
    elif outcome == "accepted":
        responded_at = offered_at + timedelta(seconds=rng.randint(20, 240))
    elif outcome == "denied":
        responded_at = offered_at + timedelta(seconds=rng.randint(15, 200))
    elif outcome == "expired":
        responded_at = expires_at
    else:  # canceled
        responded_at = offered_at + timedelta(minutes=rng.randint(3, 45))

    denial_reason: str | None = None
    if outcome == "denied":
        denial_reason = rng.choice(DENIAL_REASONS)
    elif outcome in {"expired", "canceled"} and rng.random() < 0.25:
        denial_reason = "-" if outcome == "expired" else "Canceled by client"

    amount = _amount_for(rng, service_class) if outcome == "accepted" else None

    return _Row(
        request_id=f"DR-{offered_at:%Y%m%d}-{sequence:05d}",
        client_name=client.name,
        offered_at=offered_at,
        responded_at=responded_at,
        status_source=status_source,
        outcome=outcome,
        service_type_raw=service_type_raw,
        service_class=service_class,
        denial_reason=denial_reason,
        pickup_location=rng.choice(PICKUP_LOCATIONS),
        dropoff_location=rng.choice(DROPOFF_LOCATIONS) if outcome == "accepted" else None,
        truck_assigned=rng.choice(TRUCKS) if outcome == "accepted" else None,
        driver_assigned=rng.choice(DRIVERS) if outcome == "accepted" else None,
        amount=amount,
        # THE TOWBOOK CALL NUMBER, AND IT IS BLANK ON MOST OF WHAT WE DID NOT
        # TAKE. Towbook issues one when an offer becomes a job, so in the real
        # export it is present on 99% of accepted offers, on 45% of cancelled
        # ones (accepted first, lost afterwards), and on essentially none of
        # the expired or rejected ones -- 0 of 817 Expired across the archived
        # payloads.
        #
        # It used to be filled in on every fixture row, which made the seeded
        # demo data lie about the exact thing the reports now show: it never
        # exercised the request-id fallback, so a reader of the demo would
        # expect a job number on every missed job and find none in production.
        # The 45% of cancellations is decided from `sequence`, NOT from `rng`.
        # Every other value in this fixture is drawn from that generator in a
        # fixed order, so consuming one more number here would shift the whole
        # stream and silently rewrite the fixture -- different clients,
        # different service types, and tests that assert on the shape of the
        # seeded data failing for a reason that has nothing to do with them.
        call_number=(
            f"C{100000 + sequence}"
            if outcome == "accepted" or (outcome == "canceled" and sequence % 20 < 9)
            else ""
        ),
        po_number=f"PO-{700000 + sequence}" if rng.random() < 0.6 else "",
        vehicle=rng.choice(VEHICLES),
        expires_at=expires_at,
    )


def _enforce_invariants(
    rows: list[_Row], rng: random.Random, rules: dict[str, Any]
) -> None:
    """Guarantee the signals the tests depend on, whatever the RNG did.

    Without this the suite would be flaky at the edges: a seed change or a
    tweak to the weights could quietly remove the last accepted light-service
    job and turn a real assertion into a green no-op.
    """
    # 1. unclassified rows, spread over at least two distinct raw strings.
    unclassified = [row for row in rows if row.service_class == "unclassified"]
    if len(unclassified) < MIN_UNCLASSIFIED or len({r.service_type_raw for r in unclassified}) < 2:
        candidates = [row for row in rows if row.service_class == "tow"]
        needed = max(0, MIN_UNCLASSIFIED - len(unclassified))
        for index in range(needed):
            if index >= len(candidates):
                break
            row = candidates[index]
            row.service_type_raw = UNCLASSIFIED_SERVICE_TYPES[
                index % len(UNCLASSIFIED_SERVICE_TYPES)
            ]
            row.service_class = service_class_for(row.service_type_raw, rules)

    # 2. tows that were denied or expired -- the missed_tow signal.
    missed = [
        row
        for row in rows
        if row.service_class == "tow" and row.outcome in {"denied", "expired"}
    ]
    if len(missed) < MIN_MISSED_TOWS:
        candidates = [
            row
            for row in rows
            if row.service_class == "tow" and row.outcome == "accepted"
        ]
        for index in range(MIN_MISSED_TOWS - len(missed)):
            if index >= len(candidates):
                break
            _set_outcome(candidates[index], "denied" if index % 2 == 0 else "expired", rng)

    # 3. light service accepted against policy.
    violations = [
        row
        for row in rows
        if row.service_class == "light_service" and row.outcome == "accepted"
    ]
    if len(violations) < MIN_LIGHT_SERVICE_ACCEPTED:
        candidates = [
            row
            for row in rows
            if row.service_class == "light_service" and row.outcome != "accepted"
        ]
        for index in range(MIN_LIGHT_SERVICE_ACCEPTED - len(violations)):
            if index >= len(candidates):
                break
            _set_outcome(candidates[index], "accepted", rng)

    # 4. still-live offers at the end of the window. An export taken at 23:00
    #    always has a few, and a pipeline that mishandles them inflates or
    #    deflates the acceptance rate for the newest hour.
    pending = [row for row in rows if row.outcome == "pending"]
    if len(pending) < MIN_PENDING and rows:
        last_day = max(row.offered_at.date() for row in rows)
        candidates = [
            row
            for row in rows
            if row.offered_at.date() == last_day and row.outcome != "pending"
        ]
        candidates.sort(key=lambda row: row.offered_at, reverse=True)
        for index in range(MIN_PENDING - len(pending)):
            if index >= len(candidates):
                break
            _set_outcome(candidates[index], "pending", rng)


def _set_outcome(row: _Row, outcome: str, rng: random.Random) -> None:
    """Rewrite a row's outcome and every field that depends on it."""
    row.outcome = outcome
    row.status_source = rng.choice(STATUS_STRINGS[outcome])
    if outcome == "pending":
        row.responded_at = None
        row.denial_reason = None
        row.amount = None
        row.dropoff_location = None
        row.truck_assigned = None
        row.driver_assigned = None
    elif outcome == "accepted":
        row.responded_at = row.offered_at + timedelta(seconds=rng.randint(20, 240))
        row.denial_reason = None
        row.amount = _amount_for(rng, row.service_class)
        row.dropoff_location = rng.choice(DROPOFF_LOCATIONS)
        row.truck_assigned = rng.choice(TRUCKS)
        row.driver_assigned = rng.choice(DRIVERS)
    else:
        row.amount = None
        row.dropoff_location = None
        row.truck_assigned = None
        row.driver_assigned = None
        if outcome == "denied":
            row.responded_at = row.offered_at + timedelta(seconds=rng.randint(15, 200))
            row.denial_reason = rng.choice(DENIAL_REASONS)
        elif outcome == "expired":
            row.responded_at = row.expires_at
            row.denial_reason = None
        else:
            row.responded_at = row.offered_at + timedelta(minutes=rng.randint(3, 45))
            row.denial_reason = None


def _as_dict(row: _Row) -> dict[str, Any]:
    """Convert to the canonical-field dict that :func:`write_rows_xlsx` takes."""
    return {
        "offered_at": row.offered_at,
        "request_id": row.request_id,
        "client_name": row.client_name,
        "Call Number": row.call_number,
        "PO#": row.po_number,
        "service_type_raw": row.service_type_raw,
        "Vehicle": row.vehicle,
        "Request Expiration Time": row.expires_at.strftime(_DATETIME_OUT_FORMAT),
        "status": row.status_source,
        "responded_at": (
            row.responded_at.strftime(_DATETIME_OUT_FORMAT) if row.responded_at else ""
        ),
        "denial_reason": row.denial_reason or "",
        "pickup_location": row.pickup_location,
        "dropoff_location": row.dropoff_location or "",
        "truck_assigned": row.truck_assigned or "",
        "driver_assigned": row.driver_assigned or "",
        "amount": row.amount,
        # Not written to the sheet; kept so tests can build expectations
        # without re-deriving them.
        "_outcome": row.outcome,
        "_service_class": row.service_class,
    }


# --------------------------------------------------------------------------
# Summary oracle
# --------------------------------------------------------------------------


def summarize_rows(
    rows: Iterable[dict[str, Any]],
    *,
    schema: dict[str, Any] | None = None,
    rules: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Summarise generated rows the way the metrics agent should see them.

    Computed straight from the row dicts, without touching the database or any
    pipeline code, so a test can compare the two and mean it.
    """
    schema = load_schema() if schema is None else schema
    rules = load_rules() if rules is None else rules

    total = 0
    accepted = 0
    by_status: dict[str, int] = {}
    by_class: dict[str, int] = {}
    accepted_by_class: dict[str, int] = {}
    by_day: dict[date, dict[str, int]] = {}
    by_hour: dict[datetime, dict[str, int]] = {}
    by_client: dict[str, dict[str, int]] = {}
    by_client_day: dict[tuple[date, str], dict[str, int]] = {}
    unclassified_types: set[str] = set()
    missed_tows = 0
    light_service_accepted = 0

    for row in rows:
        status = canonical_status(row.get("status", ""), schema)
        if status is None:
            continue
        offered_at = row["offered_at"]
        service_class = service_class_for(row.get("service_type_raw", ""), rules)
        client_key = (row.get("client_name") or "").strip().casefold()
        is_accepted = status == "accepted"

        total += 1
        accepted += int(is_accepted)
        by_status[status] = by_status.get(status, 0) + 1
        by_class[service_class] = by_class.get(service_class, 0) + 1
        if is_accepted:
            accepted_by_class[service_class] = accepted_by_class.get(service_class, 0) + 1

        day_bucket = by_day.setdefault(offered_at.date(), {"offered": 0, "accepted": 0})
        day_bucket["offered"] += 1
        day_bucket["accepted"] += int(is_accepted)

        hour_key = offered_at.replace(minute=0, second=0, microsecond=0)
        hour_bucket = by_hour.setdefault(hour_key, {"offered": 0, "accepted": 0})
        hour_bucket["offered"] += 1
        hour_bucket["accepted"] += int(is_accepted)

        client_bucket = by_client.setdefault(
            client_key,
            {"offered": 0, "accepted": 0, "denied": 0, "expired": 0, "canceled": 0},
        )
        client_bucket["offered"] += 1
        if status in client_bucket:
            client_bucket[status] += 1

        client_day_bucket = by_client_day.setdefault(
            (offered_at.date(), client_key),
            {"offered": 0, "accepted": 0, "denied": 0, "expired": 0, "canceled": 0},
        )
        client_day_bucket["offered"] += 1
        if status in client_day_bucket:
            client_day_bucket[status] += 1

        if service_class == "unclassified":
            unclassified_types.add(row.get("service_type_raw", ""))
        if service_class == "tow" and status in {"denied", "expired"}:
            missed_tows += 1
        if service_class == "light_service" and is_accepted:
            light_service_accepted += 1

    return {
        "offered": total,
        "accepted": accepted,
        "rate": (accepted / total) if total else None,
        "by_status": by_status,
        "by_service_class": by_class,
        "accepted_by_service_class": accepted_by_class,
        "by_day": by_day,
        "by_hour": by_hour,
        "by_client": by_client,
        "by_client_day": by_client_day,
        "unclassified_service_types": sorted(unclassified_types),
        "missed_tows": missed_tows,
        "light_service_accepted": light_service_accepted,
    }


# --------------------------------------------------------------------------
# XLSX writing
# --------------------------------------------------------------------------


def write_rows_xlsx(
    path: str | Path,
    rows: Sequence[dict[str, Any]],
    *,
    fields: Sequence[str] | None = None,
    headers: Sequence[str] | None = None,
    schema: dict[str, Any] | None = None,
    sheet_title: str = "Digital Requests",
    banner_rows: Sequence[str] = (),
) -> Path:
    """Write ``rows`` to an XLSX shaped like a Towbook export.

    ``rows`` are keyed by canonical field name (``request_id``, ``offered_at``,
    ...) or by a literal header for the documented columns that have no
    canonical field. Keys starting with ``_`` are metadata and are not written.

    ``fields``  -- explicit column order; defaults to :data:`FIELD_ORDER`
                   filtered to what the rows actually contain, then any extras.
    ``headers`` -- explicit header text, overriding the schema lookup. Used by
                   the tests that need to simulate header drift.
    ``banner_rows`` -- lines written above the header row, which is what a
                   report title looks like and what ``header_row_scan`` exists
                   to survive.
    """
    schema = load_schema() if schema is None else schema
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if fields is None:
        present: list[str] = []
        seen: set[str] = set()
        for name in FIELD_ORDER:
            if any(name in row for row in rows):
                present.append(name)
                seen.add(name)
        for row in rows:
            for name in row:
                if name.startswith("_") or name in seen:
                    continue
                present.append(name)
                seen.add(name)
        fields = present or list(FIELD_ORDER)

    if headers is None:
        headers = headers_for_fields(fields, schema)
    if len(headers) != len(fields):
        raise ValueError(
            f"headers has {len(headers)} entries but fields has {len(fields)}"
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    for line in banner_rows:
        sheet.append([line])

    sheet.append(list(headers))

    for row in rows:
        values = [row.get(name) for name in fields]
        sheet.append(values)
        written = sheet[sheet.max_row]
        for cell, value in zip(written, values):
            if isinstance(value, datetime):
                cell.number_format = _EXCEL_DATETIME_FORMAT

    # Cosmetic, and it makes the file readable when a human opens it to check
    # what the pipeline was actually given.
    for index, header in enumerate(headers, start=1):
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = max(12, min(34, len(str(header)) + 6))

    workbook.save(path)
    workbook.close()
    return path


def generate_fixture_xlsx(
    path: str | Path,
    days: int = 14,
    seed: int = 42,
    *,
    end_date: date | None = None,
    banner_rows: Sequence[str] = (),
    currency_strings: bool = True,
) -> Path:
    """Generate the realistic fixture export and return the path written.

    Deterministic: the same ``path``, ``days``, ``seed`` and ``end_date`` always
    produce the same rows. ``end_date`` defaults to today, so pass it explicitly
    from a test that asserts on specific dates.
    """
    rows = generate_rows(days=days, seed=seed, end_date=end_date)

    if currency_strings:
        # Some exports hand back a formatted currency string rather than a
        # number. Do it to a deterministic slice of the accepted rows so the
        # value_cleanup path in the ingester is genuinely exercised.
        for index, row in enumerate(rows):
            if row.get("amount") is not None and index % 7 == 0:
                row["amount"] = f"${row['amount']:,.2f}"

    return write_rows_xlsx(path, rows, banner_rows=banner_rows)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="fixture_generator",
        description="Generate a realistic Towbook Digital Requests fixture XLSX.",
    )
    parser.add_argument(
        "--out",
        "-o",
        default=str(_REPO_ROOT / "tests" / "fixtures" / "digital_requests_fixture.xlsx"),
        help="output path for the XLSX",
    )
    parser.add_argument("--days", type=int, default=14, help="number of days to generate")
    parser.add_argument("--seed", type=int, default=42, help="RNG seed (determinism)")
    parser.add_argument(
        "--end-date",
        default=None,
        help="last day of the window, YYYY-MM-DD (default: today)",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    end_date = date.fromisoformat(args.end_date) if args.end_date else None
    rows = generate_rows(days=args.days, seed=args.seed, end_date=end_date)
    summary = summarize_rows(rows)
    path = generate_fixture_xlsx(
        args.out, days=args.days, seed=args.seed, end_date=end_date
    )

    rate = summary["rate"]
    print(f"wrote {path}")
    print(f"  rows            : {summary['offered']}")
    print(f"  accepted        : {summary['accepted']} ({rate:.1%})" if rate else "  accepted: 0")
    print(f"  by status       : {summary['by_status']}")
    print(f"  by service class: {summary['by_service_class']}")
    print(f"  missed tows     : {summary['missed_tows']}")
    print(f"  light accepted  : {summary['light_service_accepted']}")
    print(f"  unclassified    : {summary['unclassified_service_types']}")
    return 0


if __name__ == "__main__":  # pragma: no cover - script entry point
    sys.exit(main())
