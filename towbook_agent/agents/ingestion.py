"""Acquired payload -> canonical request records.

Reads whatever the Acquisition agent archived, maps it onto the canonical
request record, and upserts it. Deterministic end to end: no LLM, no heuristics
beyond what config/schema.yaml declares (hard constraint #3).

TWO SOURCES, ONE PATH
---------------------
* **JSON** (``.json``) -- the verbatim API payload from
  ``agents/acquisition_api.py``. The PRIMARY source. Mapped through
  ``schema.yaml -> api:``, which is an overlay on the same schema, and keyed by
  ``callRequestId`` -- a real unique id, verified across 3,079 rows.
* **CSV / XLSX** -- the UI "Export to Excel" download from
  ``agents/acquisition.py``. Mapped through the top-level ``columns``, and keyed
  by a synthesised fingerprint, because that export's only id-shaped column
  ("Call #") is blank on exactly the unaccepted offers this system measures.

Only the *reader* differs. Header validation, column mapping, value cleanup,
datetime parsing, the status vocabulary, classification, identity and the
upsert are shared verbatim -- forking them would be forking the numbers. For a
JSON payload the "headers" are the object's keys, so header validation *is*
required-key validation, with no second implementation.

Design decisions worth knowing before you edit this file
-------------------------------------------------------

*No pandas.* openpyxl in ``read_only=True, data_only=True`` mode streams the
sheet a row at a time, which is all that is needed and keeps the dependency
tree installable on Python 3.14. JSON goes through the standard library.

*No field names in code.* Every source header and every API key lives in
config/schema.yaml, as a list of candidates per canonical field. Towbook
renaming "Service Needed" to "Service Type", or ``serviceNeeded`` to
``service``, is a one-line YAML edit (hard constraint #2).

*Header drift aborts, it does not guess.* If the required headers are not there,
the observed headers are written to config/schema.detected.yaml, a
``pipeline_failure`` event is emitted with a readable missing/unexpected/matched
diff, and **nothing is ingested**. A partial ingest silently understates the
offered count, and a wrong acceptance rate is worse than a missing one (hard
constraint #5).

*service_type_raw is immutable.* It is stored exactly as the sheet produced it,
never trimmed of meaning, and on re-ingest the first-seen value is kept rather
than overwritten (hard constraint #6). ``classifier.backfill`` re-derives
``service_class`` from it whenever the rules change.

*Idempotent.* Rows upsert on ``request_id``, so re-running the same window
yields the same table (hard constraint #4). Re-pulling is not just safe, it is
required: a request's status mutates over time (pending -> accepted, pending ->
expired), so the later pull is the more truthful one. The original first-seen
``ingested_at`` is preserved; ``source_run_id`` is updated to the run that last
saw the row.
"""

from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import re
from dataclasses import dataclass, field
from datetime import date as _date
from datetime import datetime
from datetime import time as _time
from decimal import Decimal, InvalidOperation
from itertools import chain, islice
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

import yaml
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from ..core.config_loader import get_rules, get_schema, rules_version
from ..core.db import get_session, init_db
from ..core.events import PIPELINE_FAILURE, emit_event
from ..core.models import (
    DEFAULT_ACCOUNT_ID,
    STATUS_VALUES,
    Request,
    Run,
    client_key_for,
    to_utc_naive,
    utcnow,
)
from ..core.paths import resolve_under_root
from . import classifier

__all__ = [
    "ingest",
    "IngestResult",
    "IngestionError",
    "HeaderValidationError",
    "parse_datetime",
    "map_status",
    "normalize_header",
    "api_schema",
]

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------
# Errors
# --------------------------------------------------------------------------


class IngestionError(RuntimeError):
    """Ingestion could not complete. The run is recorded as failed."""


class HeaderValidationError(IngestionError):
    """The exported headers do not match config/schema.yaml. Nothing was ingested."""

    def __init__(self, message: str, diff: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.diff: dict[str, Any] = diff or {}


# --------------------------------------------------------------------------
# Result
# --------------------------------------------------------------------------


@dataclass
class IngestResult:
    """What one ingest did. Returned to the CLI, the scheduler and Health."""

    run_id: str
    rows_read: int = 0
    rows_upserted: int = 0
    rows_inserted: int = 0
    rows_updated: int = 0
    #: raw source status string -> how many rows carried it, for statuses that
    #: config/schema.yaml has no mapping for. Surfaced on the Health view so the
    #: vocabulary can be extended without anyone having to read a log file.
    unmapped_statuses: dict[str, int] = field(default_factory=dict)
    unclassified_count: int = 0

    # -- additive detail, not part of the required contract ------------------
    rows_rejected: int = 0
    reject_reasons: dict[str, int] = field(default_factory=dict)
    unparsed_datetimes: int = 0
    unparsed_amounts: int = 0
    duplicate_request_ids: int = 0
    #: rows whose incoming service_type_raw differed from the stored one; the
    #: stored value wins, because it is immutable
    service_type_conflicts: int = 0
    header_row: int = 0
    matched_headers: dict[str, str] = field(default_factory=dict)
    unmapped_fields: list[str] = field(default_factory=list)
    #: "succeeded" or "partial"; a failure raises instead of returning
    status: str = "succeeded"
    warnings: list[str] = field(default_factory=list)
    xlsx_path: str = ""
    rules_version: str = ""

    @property
    def unmapped_status_count(self) -> int:
        return sum(self.unmapped_statuses.values())

    def as_dict(self) -> dict[str, Any]:
        """JSON-safe view, for the run log and the dashboard."""
        return {
            "run_id": self.run_id,
            "rows_read": self.rows_read,
            "rows_upserted": self.rows_upserted,
            "rows_inserted": self.rows_inserted,
            "rows_updated": self.rows_updated,
            "rows_rejected": self.rows_rejected,
            "reject_reasons": dict(self.reject_reasons),
            "unmapped_statuses": dict(self.unmapped_statuses),
            "unclassified_count": self.unclassified_count,
            "unparsed_datetimes": self.unparsed_datetimes,
            "unparsed_amounts": self.unparsed_amounts,
            "duplicate_request_ids": self.duplicate_request_ids,
            "service_type_conflicts": self.service_type_conflicts,
            "header_row": self.header_row,
            "matched_headers": dict(self.matched_headers),
            "unmapped_fields": list(self.unmapped_fields),
            "status": self.status,
            "warnings": list(self.warnings),
            "xlsx_path": self.xlsx_path,
            "rules_version": self.rules_version,
        }


# --------------------------------------------------------------------------
# Canonical field list -- the canonical request record, in column order.
# These names are fixed by the spec; every module codes against them.
# --------------------------------------------------------------------------

_TEXT_FIELDS: tuple[str, ...] = (
    "client_name",
    "denial_reason",
    "service_type_raw",
    # The customer's car, verbatim. Stored because it is the only field in the
    # feed that identifies the customer's job, and therefore the only thing
    # `duplicate_offers` can recognise a re-broadcast offer by.
    "vehicle",
    "pickup_location",
    "pickup_zip",
    "dropoff_location",
    "truck_assigned",
    "driver_assigned",
)
_DATETIME_FIELDS: tuple[str, ...] = ("offered_at", "responded_at", "expires_at")
#: Parsed with the money parser but NOT money. See Request.distance_miles.
_DECIMAL_FIELDS: tuple[str, ...] = ("amount", "distance_miles")
_CANONICAL_FIELDS: tuple[str, ...] = (
    "request_id",
    "job_number",
    "client_name",
    "vehicle",
    "offered_at",
    "responded_at",
    "status",
    "denial_reason",
    "service_type_raw",
    "pickup_location",
    "pickup_zip",
    "dropoff_location",
    "truck_assigned",
    "driver_assigned",
    "amount",
    "distance_miles",
    "expires_at",
)

#: Without these three there is no usable row: no identity, no time axis, no
#: outcome. Their absence is header drift, not a tolerable gap.
_REQUIRED_FIELDS: tuple[str, ...] = ("request_id", "offered_at", "status")

#: Every column written by the upsert, in insert order. These are TABLE COLUMN
#: NAMES, not ORM attribute names -- the insert goes through the Core -- so
#: this reads ``company_id``, the real column, and not its ``account_id``
#: synonym.
_UPSERT_COLUMNS: tuple[str, ...] = (
    "request_id",
    "company_id",
    # The Towbook job / call number. Stored, never keyed on, and NULL on most of
    # what we did not accept because Towbook does not issue one until an offer
    # becomes a job -- see models.Request.job_number.
    "job_number",
    "client_name",
    "client_key",
    "offered_at",
    "responded_at",
    "status",
    # The verbatim source status label and, when the source has one, its numeric
    # code. `status` above is the five-value controlled vocabulary; these two are
    # what let missed_work.buckets tell "Expired" from "Accept Failed" and
    # "Rejected" from "Rejected By Motor Club" at read time.
    "status_raw",
    "status_code",
    "denial_reason",
    "service_type_raw",
    "service_class",
    "vehicle",
    "pickup_location",
    "pickup_zip",
    "dropoff_location",
    "truck_assigned",
    "driver_assigned",
    "amount",
    "distance_miles",
    "expires_at",
    "ingested_at",
    "source_run_id",
)

#: Rows per statement. 24 columns x 36 rows = 864 bound parameters, under the
#: 999 limit of SQLite builds older than 3.32. Was 50 while the record had 17
#: columns; status_raw and status_code took it to 950, which is why it came down
#: to 40; job_number and vehicle have since taken the record to 24 columns. If
#: the record grows again, LOWER THIS FIRST: the failure mode is an sqlite3
#: "too many SQL variables" error at ingest, which surfaces as a
#: pipeline_failure rather than silently, but only once real data hits it.
_UPSERT_CHUNK: int = 36
#: Ids per IN (...) lookup.
_SELECT_CHUNK: int = 400

_WHITESPACE = re.compile(r"\s+")
#: Trailing decorations a portal adds to a header: "Status *", "Amount:".
_HEADER_TRAILERS = re.compile(r"[\s*:#]+$")
#: A trailing timezone abbreviation strptime cannot handle: "... 2:05 PM EDT".
_TRAILING_TZ = re.compile(r"\s+\(?[A-Z]{2,5}\)?$")

_TRUE_TOKENS = {"1", "true", "yes", "on"}

_schema_ready = False


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------


def _collapse(text: str) -> str:
    """Trim and collapse internal whitespace, including non-breaking spaces."""
    # \s does not match U+00A0 / U+202F in a str pattern, and web exports are
    # full of them; map them to a plain space before collapsing.
    text = text.replace("\u00a0", " ").replace("\u202f", " ")
    return _WHITESPACE.sub(" ", text).strip()


def normalize_header(value: Any) -> str:
    """Fold a header cell for comparison: trim, collapse, drop trailers, casefold.

    Matching is deliberately forgiving about presentation ("Date and Time  ",
    "DATE AND TIME", "Date and Time:") and completely unforgiving about
    meaning -- a header that is not in the candidate list stays unmatched.
    """
    if value is None:
        return ""
    text = _collapse(str(value))
    text = _HEADER_TRAILERS.sub("", text)
    return text.casefold()


def _as_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [value]
    if isinstance(value, (list, tuple)):
        return [str(item) for item in value if item is not None]
    return []


def _truthy(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in _TRUE_TOKENS


# --------------------------------------------------------------------------
# Timezone
# --------------------------------------------------------------------------

_tz_cache: dict[str, Any] = {}


def _source_timezone(schema: Mapping[str, Any]) -> Any:
    """Timezone the exported timestamps are written in.

    The TZ environment variable wins, then schema.yaml's ``source_timezone``.
    Everything is converted to UTC on the way into the database; local time is
    a presentation concern only.
    """
    name = (os.environ.get("TZ") or "").strip() or str(
        schema.get("source_timezone") or "America/Detroit"
    ).strip()
    cached = _tz_cache.get(name)
    if cached is not None:
        return cached
    try:
        zone: Any = ZoneInfo(name)
    except (ZoneInfoNotFoundError, ValueError, OSError) as exc:
        # Windows has no system tz database; the `tzdata` wheel supplies it. If
        # it is missing we must not silently shift every timestamp, so say so
        # loudly and fall back to UTC.
        logger.error(
            "timezone %r is unavailable (%s); falling back to UTC. "
            "Install the `tzdata` package to fix local-time windowing.",
            name,
            exc,
        )
        from datetime import timezone as _timezone

        zone = _timezone.utc
    _tz_cache[name] = zone
    return zone


# --------------------------------------------------------------------------
# Value parsing
# --------------------------------------------------------------------------


@dataclass
class _Cleanup:
    null_tokens: frozenset[str]
    collapse: bool
    verbatim_fields: frozenset[str]
    amount_strip: tuple[str, ...]
    #: Fields where a source value of 0 means "there is none". Deliberately NOT
    #: a global null token: 0 is a real answer for amount, for distance, for any
    #: count. Towbook sends `callNumber: 0` on an offer that never became a job
    #: rather than omitting the key, and a report that prints "Job #0" hands the
    #: owner a reference that finds nothing in Towbook.
    zero_means_absent: frozenset[str]

    @classmethod
    def from_schema(cls, schema: Mapping[str, Any]) -> "_Cleanup":
        block = schema.get("value_cleanup")
        if not isinstance(block, Mapping):
            block = {}
        tokens = _as_list(block.get("null_tokens")) or ["", "-", "--", "n/a", "null", "none"]
        return cls(
            null_tokens=frozenset(token.strip().casefold() for token in tokens),
            collapse=_truthy(block.get("collapse_internal_whitespace"), True),
            verbatim_fields=frozenset(
                _as_list(block.get("preserve_verbatim_fields")) or ["service_type_raw"]
            ),
            amount_strip=tuple(_as_list(block.get("amount_strip_chars")) or ["$", ",", " "]),
            zero_means_absent=frozenset(
                name.strip() for name in _as_list(block.get("zero_means_absent"))
            ),
        )

    def is_null(self, text: str) -> bool:
        return text.strip().casefold() in self.null_tokens

    def is_absent_zero(self, field: str, text: str) -> bool:
        """True when ``text`` is a zero that this field reads as "no value".

        Matches "0", "0.0" and "00" -- Towbook's JSON sends the integer 0, but a
        CSV export of the same column may well write it any of those ways, and
        a job number that is really zero does not exist.
        """
        if field not in self.zero_means_absent:
            return False
        probe = text.strip()
        if not probe:
            return False
        try:
            return float(probe) == 0.0
        except ValueError:
            return False


def _clean_text(value: Any, cleanup: _Cleanup, *, verbatim: bool) -> str | None:
    """Normalise a text cell, or return the value untouched when verbatim.

    ``verbatim`` fields (service_type_raw) are returned byte for byte. Only a
    cell with no meaning at all -- empty, whitespace, or a declared null token
    like "-" -- becomes None, because there is nothing there to preserve.
    """
    if value is None:
        return None
    if isinstance(value, str):
        text = value
    elif isinstance(value, (datetime, _date, _time)):
        text = value.isoformat()
    elif isinstance(value, bool):
        text = "true" if value else "false"
    else:
        text = str(value)

    probe = _collapse(text)
    if cleanup.is_null(probe):
        return None
    if verbatim:
        return value if isinstance(value, str) else text
    return probe if cleanup.collapse else text.strip()


def _clean_identifier(value: Any, cleanup: _Cleanup, field: str) -> str | None:
    """Normalise a cell that holds a REFERENCE rather than a number or a name.

    ``job_number`` is the case this exists for. It arrives as an integer from
    the JSON API and as a spreadsheet cell from the CSV export, and it has to
    end up as the same string either way, because it is what a human types into
    Towbook -- "125169", never "125169.0" and never "0".

    Two things happen here that :func:`_clean_text` must not do generally:

    * an integral float is rendered without its ``.0``. openpyxl hands back
      ``125169.0`` for a plain numeric cell, and a report that prints that has
      invented a reference nobody can look up.
    * a zero becomes None, for the fields schema.yaml lists under
      ``value_cleanup.zero_means_absent``. Towbook sends ``callNumber: 0`` on an
      offer that never became a job -- 1,731 of 3,124 archived records -- rather
      than omitting the key.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        # A boolean is not a reference. Nothing sends one; refuse it rather than
        # storing "true" as a job number.
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    elif isinstance(value, int):
        text = str(value)
    else:
        cleaned = _clean_text(value, cleanup, verbatim=False)
        if cleaned is None:
            return None
        text = cleaned
        # "125169.0" out of a CSV is the same reference as 125169.
        try:
            number = float(text)
        except ValueError:
            pass
        else:
            if number.is_integer():
                text = str(int(number))

    probe = _collapse(text)
    if cleanup.is_null(probe) or cleanup.is_absent_zero(field, probe):
        return None
    return probe or None


def parse_datetime(
    value: Any,
    formats: Sequence[str],
    tz: Any,
    *,
    epoch: datetime | None = None,
) -> datetime | None:
    """Parse a cell into a naive UTC datetime, or return None.

    openpyxl hands back a real ``datetime`` for a properly typed date cell, and
    those skip the format list entirely. The list exists for the case where the
    export writes timestamps as strings, which is the common shape for a web
    grid export.

    A naive value is localised into ``tz`` (the export's timezone) and then
    converted to UTC. An aware value is simply converted. Storage is always
    naive UTC -- the convention in core/models.py.
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return to_utc_naive(value if value.tzinfo else value.replace(tzinfo=tz))

    if isinstance(value, _date) and not isinstance(value, datetime):
        return to_utc_naive(datetime(value.year, value.month, value.day, tzinfo=tz))

    if isinstance(value, _time):
        return None  # a bare time has no day; it cannot anchor a window

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        # An Excel serial that data_only did not convert (the cell had no date
        # format applied). 1 = 1900-01-01, 2958465 = 9999-12-31.
        if not (1 <= float(value) <= 2958465):
            return None
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value, epoch) if epoch else from_excel(value)
        except Exception:  # pragma: no cover - depends on openpyxl internals
            return None
        if isinstance(converted, datetime):
            return to_utc_naive(converted.replace(tzinfo=tz))
        if isinstance(converted, _date):
            return to_utc_naive(datetime(converted.year, converted.month, converted.day, tzinfo=tz))
        return None

    text = _collapse(str(value))
    if not text:
        return None

    for candidate in (text, _TRAILING_TZ.sub("", text), text.replace("T", " ")):
        for pattern in formats:
            try:
                parsed = datetime.strptime(candidate, pattern)
            except (ValueError, TypeError):
                continue
            return to_utc_naive(parsed if parsed.tzinfo else parsed.replace(tzinfo=tz))

    # Last resort: ISO 8601, which strptime patterns often miss because of the
    # optional fractional seconds and the trailing Z.
    iso = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(iso)
    except ValueError:
        return None
    return to_utc_naive(parsed if parsed.tzinfo else parsed.replace(tzinfo=tz))


def _parse_amount(value: Any, cleanup: _Cleanup) -> tuple[Decimal | None, bool]:
    """Return (amount, failed). ``failed`` is True only for unparseable text."""
    if value is None or isinstance(value, bool):
        return None, False
    if isinstance(value, Decimal):
        return _quantize(value), False
    if isinstance(value, (int, float)):
        return _quantize(Decimal(str(value))), False

    text = str(value)
    if cleanup.is_null(_collapse(text)):
        return None, False
    for token in cleanup.amount_strip:
        if token:
            text = text.replace(token, "")
    text = _collapse(text)
    negative = False
    if text.startswith("(") and text.endswith(")"):  # accounting negative
        negative, text = True, text[1:-1]
    if text.endswith("-"):
        negative, text = True, text[:-1]
    if not text:
        return None, False
    try:
        amount = Decimal(text)
    except (InvalidOperation, ValueError):
        return None, True
    return _quantize(-amount if negative else amount), False


def _quantize(value: Decimal) -> Decimal:
    try:
        return value.quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):  # pragma: no cover - absurd magnitudes
        return value


def _split_source_status(raw: Any) -> tuple[str | None, int | None]:
    """Split a source status cell into ``(verbatim label, numeric code)``.

    The CSV export writes a label and no code ("Goa Approved By Motor Club").
    The JSON API carries both, and ``schema.yaml -> api.columns.status`` prefers
    ``statusName``, falling back to the numeric ``status`` when the label is
    absent. So a cell is one or the other, never both, and this reports which.

    A purely numeric cell yields ``(None, code)``: storing "40" as a label would
    be storing a code in a text column and inviting somebody to string-match on
    it later.
    """
    if raw is None:
        return (None, None)
    if isinstance(raw, bool):  # bool is an int; a boolean status is nonsense
        return (_collapse(str(raw)) or None, None)
    if isinstance(raw, int):
        return (None, int(raw))
    if isinstance(raw, float) and float(raw).is_integer():
        return (None, int(raw))

    text = _collapse(str(raw))
    if not text:
        return (None, None)
    stripped = text.strip()
    if stripped.isdigit() or (stripped[:1] == "-" and stripped[1:].isdigit()):
        try:
            return (None, int(stripped))
        except ValueError:  # pragma: no cover - isdigit already guarantees this
            return (text, None)
    return (text, None)


def map_status(raw: Any, vocabulary: Mapping[str, str]) -> str | None:
    """Map a source status onto the controlled vocabulary, or None if unknown.

    The indirection through schema.yaml is what lets denied / expired /
    canceled be split apart later with no migration: the stored value changes,
    the schema does not.
    """
    key = _collapse("" if raw is None else str(raw)).casefold()
    if not key:
        return None
    mapped = vocabulary.get(key)
    if mapped is None:
        return None
    mapped = str(mapped).strip().casefold()
    if mapped not in STATUS_VALUES:
        logger.warning(
            "schema.yaml maps status %r to %r, which is not in the controlled "
            "vocabulary %s; treating the row as unmapped",
            key,
            mapped,
            ", ".join(STATUS_VALUES),
        )
        return None
    return mapped


# --------------------------------------------------------------------------
# Header validation
# --------------------------------------------------------------------------


def _build_header_index(headers: Sequence[Any]) -> dict[str, int]:
    """normalised header -> column index. First occurrence wins on duplicates."""
    index: dict[str, int] = {}
    for position, value in enumerate(headers):
        key = normalize_header(value)
        if key and key not in index:
            index[key] = position
    return index


def _score_header_row(headers: Sequence[Any], required: Sequence[str]) -> int:
    index = _build_header_index(headers)
    return sum(1 for name in required if normalize_header(name) in index)


def _find_header_row(
    scanned: Sequence[Sequence[Any]], schema: Mapping[str, Any]
) -> tuple[int, list[Any]]:
    """Locate the header row inside the first N rows of the sheet.

    Scanning rather than trusting ``header_row: 1`` survives an export with a
    title or a logo banner above the grid, which is common for a web report.
    Returns (0-based index within ``scanned``, the row).
    """
    sheet_cfg = schema.get("sheet") if isinstance(schema.get("sheet"), Mapping) else {}
    required = _as_list(schema.get("required_headers"))
    declared = int(sheet_cfg.get("header_row") or 1)

    if not _truthy(sheet_cfg.get("header_row_scan"), True):
        position = max(0, declared - 1)
        row = list(scanned[position]) if position < len(scanned) else []
        return position, row

    best_position, best_score = -1, -1
    for position, row in enumerate(scanned):
        score = _score_header_row(row, required) if required else 0
        if not required:
            # Nothing to score against: fall back to the first row that has any
            # non-empty cell.
            score = 1 if any(normalize_header(cell) for cell in row) else 0
        if score > best_score:
            best_position, best_score = position, score
        if required and score == len(required):
            break

    if best_position < 0:
        return 0, list(scanned[0]) if scanned else []
    return best_position, list(scanned[best_position])


def _map_columns(
    header_index: Mapping[str, int], schema: Mapping[str, Any]
) -> tuple[dict[str, int], dict[str, str], list[str]]:
    """Resolve canonical field -> column index using schema.yaml candidate lists.

    Returns (column_map, matched {canonical: source header}, unmapped fields).
    Candidate order is priority order.
    """
    columns = schema.get("columns")
    if not isinstance(columns, Mapping):
        raise HeaderValidationError("config/schema.yaml has no `columns` mapping")

    column_map: dict[str, int] = {}
    matched: dict[str, str] = {}
    unmapped: list[str] = []

    for field_name in _CANONICAL_FIELDS:
        candidates = _as_list(columns.get(field_name))
        for candidate in candidates:
            key = normalize_header(candidate)
            if key in header_index:
                column_map[field_name] = header_index[key]
                matched[field_name] = candidate
                break
        else:
            unmapped.append(field_name)

    return column_map, matched, unmapped


def _status_source_indexes(
    header_index: Mapping[str, int], schema: Mapping[str, Any]
) -> tuple[int, ...]:
    """Indexes of EVERY declared source column for ``status``, in schema order.

    ``_map_columns`` keeps only the winning candidate per canonical field, which
    is right for every field but this one. ``schema.yaml -> api.columns.status``
    declares ``[statusName, status]``: the JSON API sends BOTH, the label wins
    the mapping, and the numeric code -- which ``requests.status_code`` exists to
    hold and which ``rules.yaml -> missed_work.buckets`` keys its whole 14-entry
    table on -- was being dropped on the floor. Measured on a real 2,932-record
    pull: ``status_code`` NULL on 2,932 of 2,932 rows, so the numeric bucket map
    never once fired and every row fell through to the label map.

    That still produced correct buckets, because the labels are exact. But it
    left the exact-integer route dead and the system resting on strings Towbook
    owns and can reword -- and a reworded label does not fail loudly, it falls
    through to the coarse canonical map, where "Accept Failed" silently becomes
    ``no_response`` instead of ``accept_failed``.

    So: collect every declared source column that the payload actually has, and
    let :func:`_build_record` take the label from one and the code from another.
    The candidate list stays in schema.yaml; no source column name is named here.
    """
    columns = schema.get("columns")
    if not isinstance(columns, Mapping):
        return ()
    indexes: list[int] = []
    for candidate in _as_list(columns.get("status")):
        index = header_index.get(normalize_header(candidate))
        if index is not None and index not in indexes:
            indexes.append(index)
    return tuple(indexes)


def _validate_headers(
    headers: Sequence[Any],
    header_row_number: int,
    schema: Mapping[str, Any],
    xlsx_path: Path,
    sheet_name: str,
    source_format: str = "xlsx",
) -> tuple[dict[str, int], dict[str, str], list[str], dict[str, Any]]:
    """Validate the header row, or write schema.detected.yaml and abort.

    Hard constraint #5 in practice: a header change is a silent data-loss bug
    if you let it through, so it fails the run instead. The observed headers
    are written to disk so reconciling schema.yaml is a copy/paste, not an
    archaeology exercise.

    ``source_format`` only changes the wording of the diagnostic. For a JSON
    payload the "headers" are the payload's KEYS and ``required_headers`` came
    from ``api.required_keys``, so **header validation for JSON is required-key
    validation** -- the same check, the same abort, the same
    schema.detected.yaml, with no second implementation to keep in step.
    """
    validation = schema.get("header_validation")
    if not isinstance(validation, Mapping):
        validation = {}

    observed = [str(cell).strip() for cell in headers if normalize_header(cell)]
    header_index = _build_header_index(headers)
    column_map, matched, unmapped = _map_columns(header_index, schema)

    required = _as_list(schema.get("required_headers"))
    missing_required = [name for name in required if normalize_header(name) not in header_index]

    # Which observed headers does schema.yaml know nothing about?
    known: set[str] = set()
    columns = schema.get("columns") if isinstance(schema.get("columns"), Mapping) else {}
    for candidates in columns.values():
        known.update(normalize_header(item) for item in _as_list(candidates))
    known.update(normalize_header(item) for item in _as_list(schema.get("optional_columns")))
    known.update(normalize_header(item) for item in required)
    unexpected = [name for name in observed if normalize_header(name) not in known]

    # request_id stops being a required *column* when schema.yaml configures a
    # deterministic fallback identity: the verified Towbook export has no id
    # column at all, and that is the export's shape, not header drift.
    identity = schema.get("identity") if isinstance(schema.get("identity"), Mapping) else {}
    has_fallback_identity = bool(
        _as_list(identity.get("fallback_fields")) or _as_list(identity.get("fallback_source_columns"))
    )
    required_fields = [
        name
        for name in _REQUIRED_FIELDS
        if not (name == "request_id" and has_fallback_identity)
    ]
    missing_fields = [name for name in required_fields if name not in column_map]

    diff: dict[str, Any] = {
        "xlsx_path": str(xlsx_path),
        "sheet": sheet_name,
        "source_format": source_format,
        "header_row": header_row_number,
        "observed_headers": observed,
        "matched": matched,
        "missing_required_headers": missing_required,
        "unmapped_canonical_fields": unmapped,
        "missing_required_fields": missing_fields,
        "unexpected_headers": unexpected,
    }

    strict = str(validation.get("mode") or "strict").strip().lower() == "strict"
    allow_extra = _truthy(validation.get("allow_extra_headers"), True)
    fatal = bool(missing_required or missing_fields) or (unexpected and not allow_extra)

    if not fatal:
        if unexpected:
            logger.info(
                "ingest saw %d unmapped extra column(s), which is allowed: %s",
                len(unexpected),
                ", ".join(unexpected),
            )
        return column_map, matched, unmapped, diff

    detected_path = _write_detected_schema(schema, diff)
    diff["detected_output"] = str(detected_path)
    message = _format_header_diff(diff)

    if not strict:
        # Lenient mode still reports, but only aborts when the row is unusable.
        logger.warning("header validation is not strict; continuing despite drift.\n%s", message)
        if not missing_fields:
            return column_map, matched, unmapped, diff

    raise HeaderValidationError(message, diff)


def _format_header_diff(diff: Mapping[str, Any]) -> str:
    """A diff a human can act on without opening the spreadsheet."""
    if diff.get("source_format") == "json":
        lines = [
            "Towbook API payload keys do not match config/schema.yaml `api:`. "
            "NOTHING WAS INGESTED.",
            f"  file          : {diff.get('xlsx_path')}",
            f"  payload       : JSON, keys read from the records themselves",
        ]
    else:
        lines = [
            "Request Log export headers do not match config/schema.yaml. "
            "NOTHING WAS INGESTED.",
            f"  file          : {diff.get('xlsx_path')}",
            f"  sheet         : {diff.get('sheet')}",
            f"  header row    : {diff.get('header_row')}",
        ]
    matched = diff.get("matched") or {}
    if matched:
        lines.append("  matched       :")
        for canonical, source in matched.items():
            lines.append(f"      {canonical:<17} <- {source!r}")
    for label, key in (
        ("missing (required headers)", "missing_required_headers"),
        ("missing (required fields) ", "missing_required_fields"),
        ("unmapped canonical fields ", "unmapped_canonical_fields"),
        ("unexpected headers        ", "unexpected_headers"),
    ):
        values = diff.get(key) or []
        if values:
            lines.append(f"  {label}: {', '.join(str(v) for v in values)}")
    lines.append(f"  observed      : {', '.join(str(v) for v in diff.get('observed_headers') or [])}")
    if diff.get("detected_output"):
        lines.append(f"  written to    : {diff['detected_output']}")
        lines.append(
            "  Reconcile config/schema.yaml against that file, then re-run the same "
            "window -- ingestion is idempotent."
        )
    return "\n".join(lines)


def _write_detected_schema(schema: Mapping[str, Any], diff: Mapping[str, Any]) -> Path:
    """Write what we actually saw, so schema.yaml can be reconciled by hand.

    Never written on a successful run: its presence means "a human owes this
    pipeline five minutes".
    """
    validation = schema.get("header_validation")
    target = "config/schema.detected.yaml"
    if isinstance(validation, Mapping) and validation.get("detected_output"):
        target = str(validation["detected_output"])
    path = resolve_under_root(target)

    document = {
        "detected_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_file": diff.get("xlsx_path"),
        "source_format": diff.get("source_format"),
        "sheet": diff.get("sheet"),
        "header_row": diff.get("header_row"),
        "observed_headers": list(diff.get("observed_headers") or []),
        "matched": dict(diff.get("matched") or {}),
        "missing_required_headers": list(diff.get("missing_required_headers") or []),
        "missing_required_fields": list(diff.get("missing_required_fields") or []),
        "unmapped_canonical_fields": list(diff.get("unmapped_canonical_fields") or []),
        "unexpected_headers": list(diff.get("unexpected_headers") or []),
    }
    banner = (
        "# AUTO-GENERATED by towbook_agent.agents.ingestion -- do not edit.\n"
        "# The Request Log export did not match config/schema.yaml, so the run\n"
        "# was aborted before storing anything. This file records exactly what\n"
        "# the sheet contained. Copy the real header text into the candidate\n"
        "# lists in config/schema.yaml and re-run the window.\n"
    )
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            banner + yaml.safe_dump(document, sort_keys=False, allow_unicode=True),
            encoding="utf-8",
        )
        logger.error("wrote observed headers to %s", path)
    except OSError as exc:  # pragma: no cover - disk level failure
        logger.error("could not write %s: %s", path, exc)
    return path


# --------------------------------------------------------------------------
# Sheet parsing
# --------------------------------------------------------------------------


@dataclass
class _Parsed:
    records: list[dict[str, Any]] = field(default_factory=list)
    rows_read: int = 0
    rows_rejected: int = 0
    reject_reasons: dict[str, int] = field(default_factory=dict)
    unmapped_statuses: dict[str, int] = field(default_factory=dict)
    unparsed_datetimes: int = 0
    unparsed_amounts: int = 0
    duplicate_request_ids: int = 0
    header_row: int = 0
    matched: dict[str, str] = field(default_factory=dict)
    unmapped_fields: list[str] = field(default_factory=list)
    sheet_name: str = ""


def _select_sheet(workbook: Any, schema: Mapping[str, Any]) -> Any:
    sheet_cfg = schema.get("sheet") if isinstance(schema.get("sheet"), Mapping) else {}
    wanted = sheet_cfg.get("name")
    if wanted:
        wanted = str(wanted)
        for name in workbook.sheetnames:
            if name.strip().casefold() == wanted.strip().casefold():
                return workbook[name]
        raise IngestionError(
            f"worksheet {wanted!r} is not in the workbook; sheets present: "
            f"{', '.join(workbook.sheetnames)}"
        )
    return workbook.active if workbook.active is not None else workbook[workbook.sheetnames[0]]


def _tabular_format(path: Path, sheet_cfg: Mapping[str, Any]) -> str:
    """Return ``"json"``, ``"csv"`` or ``"xlsx"`` for ``path``, driven by schema.yaml.

    Three real shapes reach this function:

    * ``json`` -- the archive written by ``agents.acquisition_api``, which is
      the PRIMARY source. Verbatim API payloads, keyed by ``callRequestId``.
    * ``csv``  -- Towbook's "Export to Excel" actually delivers a CSV (verified
      live): comma separated, CRLF, two preamble lines above the header row.
    * ``xlsx`` -- what the menu item claims to produce, and what the test
      fixtures use. Handled unchanged if Towbook ever switches.

    openpyxl cannot open a CSV and neither reader can open JSON, which is why
    this dispatch exists rather than a single loader. Everything downstream of
    the read is shared.
    """
    suffix = path.suffix.lower()
    json_suffixes = {
        str(item).strip().lower()
        for item in (_as_list(sheet_cfg.get("json_suffixes")) or [".json"])
    }
    if suffix in json_suffixes:
        return "json"
    csv_suffixes = {
        str(item).strip().lower()
        for item in (_as_list(sheet_cfg.get("csv_suffixes")) or [".csv", ".tsv", ".txt"])
    }
    return "csv" if suffix in csv_suffixes else "xlsx"


# --------------------------------------------------------------------------
# JSON API payloads -- the primary source
# --------------------------------------------------------------------------


def api_schema(schema: Mapping[str, Any]) -> Mapping[str, Any]:
    """Overlay ``schema["api"]`` onto the base schema for a JSON payload.

    Deliberately an overlay rather than a second schema. Only the source-shaped
    keys change -- which JSON key feeds which canonical field, which keys are
    required, and whether there is a fallback identity. ``status_vocabulary``,
    ``datetime_formats``, ``value_cleanup``, ``source_timezone`` and
    ``unknown_status_action`` stay exactly as they are, so the CSV path and the
    API path share one definition of what a canonical request record is and one
    implementation of how a value is cleaned.

    Returns ``schema`` untouched when there is no ``api`` block, so a config
    written before the API path existed still ingests JSON on a best-effort
    basis rather than crashing.
    """
    block = schema.get("api")
    if not isinstance(block, Mapping):
        logger.warning(
            "config/schema.yaml has no `api:` block; falling back to the CSV column "
            "candidates for a JSON payload, which will almost certainly fail header "
            "validation. Add the api block."
        )
        return schema

    effective = dict(schema)
    if isinstance(block.get("columns"), Mapping):
        effective["columns"] = block["columns"]
    if block.get("required_keys") is not None:
        effective["required_headers"] = _as_list(block.get("required_keys"))
    if block.get("optional_keys") is not None:
        effective["optional_columns"] = _as_list(block.get("optional_keys"))
    if "identity" in block:
        # `identity: {}` is meaningful -- it says "there is no fallback, a
        # record without callRequestId is a defect" -- so it must survive the
        # falsy check a plain `or` would apply.
        effective["identity"] = block.get("identity") or {}

    extra = block.get("status_vocabulary_extra")
    if isinstance(extra, Mapping):
        merged = dict(schema.get("status_vocabulary") or {})
        merged.update(extra)
        effective["status_vocabulary"] = merged

    return effective


def _json_records(payload: Any, sheet_cfg: Mapping[str, Any], path: Path) -> list[Mapping[str, Any]]:
    """Pull the record list out of an archived API payload.

    Handles the envelope ``acquisition_api`` writes (``pages: [{records: [...]}]``,
    one entry per API page, each holding the verbatim response array), a plain
    ``{"records": [...]}`` wrapper, and a bare top-level array -- so a response
    saved by hand with curl ingests without any repackaging.
    """
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, Mapping)]

    if not isinstance(payload, Mapping):
        raise IngestionError(
            f"{path} contains a JSON {type(payload).__name__}; expected an array of call "
            "requests or an object wrapping one."
        )

    for key in _as_list(sheet_cfg.get("json_page_keys")) or ["pages"]:
        pages = payload.get(key)
        if isinstance(pages, list) and pages:
            records: list[Mapping[str, Any]] = []
            for page in pages:
                if not isinstance(page, Mapping):
                    continue
                for inner in _as_list(sheet_cfg.get("json_record_keys")) or ["records"]:
                    batch = page.get(inner)
                    if isinstance(batch, list):
                        records.extend(item for item in batch if isinstance(item, Mapping))
                        break
            return records

    for key in _as_list(sheet_cfg.get("json_record_keys")) or ["records", "data", "items"]:
        batch = payload.get(key)
        if isinstance(batch, list):
            return [item for item in batch if isinstance(item, Mapping)]

    raise IngestionError(
        f"{path} is a JSON object with no record list. Looked for "
        f"{', '.join(_as_list(sheet_cfg.get('json_page_keys')) or ['pages'])} and "
        f"{', '.join(_as_list(sheet_cfg.get('json_record_keys')) or ['records'])}; "
        f"top-level keys present: {', '.join(sorted(str(k) for k in payload))}."
    )


def _read_json_rows(
    path: Path, sheet_cfg: Mapping[str, Any]
) -> tuple[list[str], list[list[Any]]]:
    """Read an archived API payload into ``(keys, rows)``.

    The returned pair is shaped exactly like a header row plus ``values_only``
    data rows, which is what lets the entire validation, mapping, normalisation
    and upsert path below be shared verbatim with the CSV and XLSX readers
    instead of forked.

    ``keys`` is the union of every key seen across the records, in first-seen
    order. A field that is absent from some records still gets a column, and a
    field Towbook adds shows up in the validator's "unexpected" list rather
    than vanishing.
    """
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise IngestionError(f"{path} is not valid JSON: {exc}") from exc
    except OSError as exc:
        raise IngestionError(f"could not open {path}: {type(exc).__name__}: {exc}") from exc

    records = _json_records(payload, sheet_cfg, path)

    keys: list[str] = []
    seen: set[str] = set()
    for record in records:
        for key in record:
            name = str(key)
            if name not in seen:
                seen.add(name)
                keys.append(name)

    rows = [[record.get(key) for key in keys] for record in records]
    return keys, rows


def _read_csv_rows(path: Path, sheet_cfg: Mapping[str, Any]) -> list[list[Any]]:
    """Read a CSV export into rows shaped exactly like openpyxl's values_only.

    The preamble lines above the header are deliberately NOT skipped by a fixed
    offset: the existing header scan finds the real header row, which also
    survives the portal changing how many preamble lines it writes.
    """
    encodings = [
        str(item)
        for item in (_as_list(sheet_cfg.get("csv_encodings")) or ["utf-8-sig", "cp1252", "latin-1"])
    ]
    configured = sheet_cfg.get("csv_delimiter")
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                if configured:
                    delimiter = str(configured)
                else:
                    sample = handle.read(8192)
                    handle.seek(0)
                    try:
                        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
                    except csv.Error:
                        delimiter = ","
                return [list(row) for row in csv.reader(handle, delimiter=delimiter)]
        except (UnicodeDecodeError, LookupError) as exc:
            last_error = exc
            continue
        except OSError as exc:
            raise IngestionError(f"could not open {path}: {type(exc).__name__}: {exc}") from exc

    raise IngestionError(
        f"could not decode {path} as text using any of {', '.join(encodings)}: {last_error}"
    )


def _parse_sheet(
    xlsx_path: Path,
    schema: Mapping[str, Any],
    rules: Mapping[str, Any],
    run_id: str,
    company_id: str,
) -> _Parsed:
    """Stream the payload and build canonical records. Reads, never writes.

    Three source formats, ONE record-building path. The reader is chosen from
    the archived file's own format -- JSON from the API, CSV or XLSX from the
    UI export -- and each one produces the same ``(headers, rows)`` shape. For
    a JSON payload the "headers" are the object's keys, which is what makes
    header validation and required-key validation the same code.

    Everything after the read -- header validation, column mapping, value
    cleanup, datetime parsing, the status vocabulary, classification, identity
    and the upsert -- is shared verbatim. A second normalisation path would be
    a second set of numbers.
    """
    sheet_cfg = schema.get("sheet") if isinstance(schema.get("sheet"), Mapping) else {}
    source_format = _tabular_format(xlsx_path, sheet_cfg)
    if source_format == "json":
        # The api: overlay swaps in the JSON field names, the JSON required
        # keys and the (absent) fallback identity. Nothing else changes.
        schema = api_schema(schema)
    scan_limit = max(1, int(sheet_cfg.get("header_row_scan_limit") or 10))
    formats = _as_list(schema.get("datetime_formats"))
    cleanup = _Cleanup.from_schema(schema)
    tz = _source_timezone(schema)
    vocabulary = {
        _collapse(str(key)).casefold(): str(value)
        for key, value in (schema.get("status_vocabulary") or {}).items()
    }
    unknown_action = str(schema.get("unknown_status_action") or "pending").strip().lower()
    identity = schema.get("identity") if isinstance(schema.get("identity"), Mapping) else {}
    default_class = classifier.default_service_class(rules)
    ingested_at = utcnow()

    parsed = _Parsed()

    workbook = None
    rows: Iterator[Sequence[Any]]
    epoch: datetime | None = None
    #: For JSON the key list IS the header row, so there is nothing to scan for
    #: and no preamble to skip. None means "scan, as for a sheet".
    json_headers: list[str] | None = None

    if source_format == "json":
        # The primary source: the verbatim API payload archived by
        # agents/acquisition_api.py, keyed by callRequestId.
        json_headers, all_rows = _read_json_rows(xlsx_path, sheet_cfg)
        parsed.sheet_name = xlsx_path.name
        rows = iter(all_rows)
    elif source_format == "csv":
        # Verified real Towbook export: CSV, whatever the menu item is called.
        all_rows = _read_csv_rows(xlsx_path, sheet_cfg)
        parsed.sheet_name = xlsx_path.name
        rows = iter(all_rows)
    else:
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        except IngestionError:
            raise
        except Exception as exc:
            raise IngestionError(
                f"could not open {xlsx_path}: {type(exc).__name__}: {exc}"
            ) from exc
        sheet = _select_sheet(workbook, schema)
        parsed.sheet_name = getattr(sheet, "title", "") or ""
        epoch = getattr(workbook, "epoch", None)
        rows = sheet.iter_rows(values_only=True)

    try:
        if json_headers is not None:
            # A JSON payload has no header row to find: the keys are the
            # header. An empty record list is NOT header drift -- a genuinely
            # quiet window returns [] -- so it is reported as zero rows rather
            # than aborting the run, which the sheet path cannot distinguish.
            headers: list[Any] = list(json_headers)
            scanned = []
            position = -1
            parsed.header_row = 0
            if not headers:
                logger.warning(
                    "%s contains no records at all. If the window really was quiet this is "
                    "correct; if it was not, check the acquisition run's record_count.",
                    xlsx_path,
                )
        else:
            scanned = [list(row) for row in islice(rows, scan_limit)]
            if not scanned:
                raise HeaderValidationError(
                    f"{xlsx_path} contains no rows at all -- the export produced an empty sheet.",
                    {
                        "xlsx_path": str(xlsx_path),
                        "sheet": parsed.sheet_name,
                        "observed_headers": [],
                    },
                )

            position, headers = _find_header_row(scanned, schema)
            parsed.header_row = position + 1

        if json_headers is not None and not headers:
            # Nothing to validate and nothing to map; fall through with an
            # empty result rather than reporting every canonical field missing.
            parsed.matched = {}
            parsed.unmapped_fields = list(_CANONICAL_FIELDS)
            return parsed

        column_map, matched, unmapped_fields, _diff = _validate_headers(
            headers,
            parsed.header_row,
            schema,
            xlsx_path,
            parsed.sheet_name,
            source_format=source_format,
        )
        parsed.matched = matched
        parsed.unmapped_fields = unmapped_fields
        # Lets the identity fingerprint reach source columns that have no
        # canonical field of their own.
        header_index = _build_header_index(headers)
        status_columns = _status_source_indexes(header_index, schema)

        seen: dict[str, int] = {}
        data_rows: Iterable[Sequence[Any]] = chain(scanned[position + 1 :], rows)

        for row in data_rows:
            if row is None or not any(cell is not None and str(cell).strip() for cell in row):
                continue  # blank spacer row, or the grid's trailing padding
            parsed.rows_read += 1

            record, problem = _build_record(
                row,
                column_map=column_map,
                cleanup=cleanup,
                formats=formats,
                tz=tz,
                epoch=epoch,
                vocabulary=vocabulary,
                unknown_action=unknown_action,
                rules=rules,
                default_class=default_class,
                company_id=company_id,
                run_id=run_id,
                ingested_at=ingested_at,
                parsed=parsed,
                identity=identity,
                header_index=header_index,
                status_columns=status_columns,
            )
            if problem is not None:
                parsed.rows_rejected += 1
                parsed.reject_reasons[problem] = parsed.reject_reasons.get(problem, 0) + 1
                continue

            request_id = record["request_id"]
            if request_id in seen:
                # The same request twice in one export: the later row is the
                # more recent truth, so it wins.
                parsed.duplicate_request_ids += 1
                parsed.records[seen[request_id]] = record
            else:
                seen[request_id] = len(parsed.records)
                parsed.records.append(record)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:  # pragma: no cover - openpyxl close is best effort
                pass

    return parsed


def _cell(row: Sequence[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _build_record(
    row: Sequence[Any],
    *,
    column_map: Mapping[str, int],
    cleanup: _Cleanup,
    formats: Sequence[str],
    tz: Any,
    epoch: datetime | None,
    vocabulary: Mapping[str, str],
    unknown_action: str,
    rules: Mapping[str, Any],
    default_class: str,
    company_id: str,
    run_id: str,
    ingested_at: datetime,
    parsed: _Parsed,
    identity: Mapping[str, Any],
    header_index: Mapping[str, int],
    status_columns: Sequence[int] = (),
) -> tuple[dict[str, Any], str | None]:
    """Turn one sheet row into a canonical record, or explain why it cannot be."""
    record: dict[str, Any] = {name: None for name in _UPSERT_COLUMNS}
    # Which towing company was offered this job. Stamped on every row at the
    # moment of ingest, because it is the only point in the system that still
    # knows which login the payload was pulled with.
    record["company_id"] = company_id
    record["source_run_id"] = run_id
    record["ingested_at"] = ingested_at

    raw_id = _clean_text(_cell(row, column_map.get("request_id")), cleanup, verbatim=False)

    for name in _TEXT_FIELDS:
        record[name] = _clean_text(
            _cell(row, column_map.get(name)),
            cleanup,
            verbatim=name in cleanup.verbatim_fields,
        )

    record["client_key"] = client_key_for(record["client_name"])

    # The Towbook job / call number, if this offer ever got one. It is a
    # reference, not text and not a quantity, so it goes through
    # _clean_identifier: no ".0" tail, and a 0 from the source becomes NULL
    # rather than the job number zero. Blank on most unaccepted offers by
    # design -- see models.Request.job_number.
    record["job_number"] = _clean_identifier(
        _cell(row, column_map.get("job_number")), cleanup, "job_number"
    )

    for name in _DATETIME_FIELDS:
        index = column_map.get(name)
        raw = _cell(row, index)
        value = parse_datetime(raw, formats, tz, epoch=epoch)
        if value is None and index is not None and raw is not None and str(raw).strip():
            parsed.unparsed_datetimes += 1
        record[name] = value

    for name in _DECIMAL_FIELDS:
        value, failed = _parse_amount(_cell(row, column_map.get(name)), cleanup)
        record[name] = value
        if failed:
            # Counted together: both are "a number the source sent that we could
            # not read", and the reject report names the field.
            parsed.unparsed_amounts += 1

    raw_status = _cell(row, column_map.get("status"))
    status = map_status(raw_status, vocabulary)
    if status is None:
        key = _collapse("" if raw_status is None else str(raw_status)) or "(blank)"
        parsed.unmapped_statuses[key] = parsed.unmapped_statuses.get(key, 0) + 1
        if unknown_action == "reject_row":
            return record, "unknown_status"
        if unknown_action == "ignore":
            status = None
        else:
            # Default, and what the build spec calls for: keep the row, park it
            # in `pending`, and report the unmapped string on the Health view so
            # schema.yaml can be extended. Dropping the row would understate the
            # offered count, which is the number the whole system exists to get
            # right.
            status = "pending"
    record["status"] = status

    # Keep the source status alongside the collapsed one. `status` is a
    # five-value controlled vocabulary and the portal emits fourteen statuses;
    # collapsing them merges "Expired" (nobody answered) with "Accept Failed"
    # (we answered and it failed), and separates "Rejected" (we said no) from
    # "Rejected By Motor Club" (the client pulled it) by nothing at all. The
    # missed-work model needs those distinctions, and storing the verbatim label
    # is what lets rules.yaml -> missed_work.buckets re-cut them at read time
    # with no migration -- the same payoff service_type_raw already delivers.
    #
    # Unlike service_type_raw this is NOT immutable across re-ingest: an offer's
    # status legitimately changes (pending -> accepted, pending -> expired), so
    # the later pull is the more truthful one and is allowed to win.
    record["status_raw"], record["status_code"] = _split_source_status(raw_status)

    # The mapped cell is one or the other, never both. The JSON API sends both,
    # in two different columns, so fill whichever half is still missing from the
    # other declared source column -- see _status_source_indexes for why the
    # numeric code is worth going back for. First non-None wins, in schema order;
    # the mapped column has already had its say, so this can only add.
    if record["status_raw"] is None or record["status_code"] is None:
        for index in status_columns:
            label, code = _split_source_status(_cell(row, index))
            if record["status_raw"] is None and label is not None:
                record["status_raw"] = label
            if record["status_code"] is None and code is not None:
                record["status_code"] = code
            if record["status_raw"] is not None and record["status_code"] is not None:
                break

    record["service_class"] = classifier.classify_service(record["service_type_raw"] or "", rules)

    # Identity last, because the fallback fingerprint is built from the fields
    # parsed above. The verified Towbook export has no id column at all and
    # leaves "Call #" blank on every offer that was never accepted -- which is
    # precisely the population this system exists to measure. Rejecting those
    # rows would delete the numerator's complement from the denominator.
    record["request_id"] = raw_id or _fallback_request_id(record, row, header_index, identity)
    if not record["request_id"]:
        return record, "missing_request_id"
    return record, None


def _fallback_request_id(
    record: Mapping[str, Any],
    row: Sequence[Any],
    header_index: Mapping[str, int],
    identity: Mapping[str, Any],
) -> str | None:
    """A deterministic surrogate key for an export row that carries no id.

    Built from the row's own content, so the same offer seen in a later
    overlapping export produces the same key and upserts instead of
    duplicating -- hard constraint #4. Returns None when the configured fields
    are all empty, because a fingerprint of nothing would collapse every blank
    row onto one id.
    """
    fields = _as_list(identity.get("fallback_fields"))
    source_columns = _as_list(identity.get("fallback_source_columns"))
    if not fields and not source_columns:
        return None

    parts: list[str] = []
    for name in fields:
        value = record.get(str(name))
        if value is None or value == "":
            parts.append("")
            continue
        if isinstance(value, datetime):
            parts.append(value.replace(microsecond=0).isoformat())
        else:
            parts.append(_collapse(str(value)).casefold())

    # Raw source columns carry entropy the canonical record has nowhere to put
    # (the real export's "Expiration Date" and "Vehicle"). Only offer-time
    # columns belong here: anything that changes once the job is accepted would
    # give the same offer two different keys and double-count it.
    for header in source_columns:
        cell = _cell(row, header_index.get(normalize_header(header)))
        parts.append("" if cell is None else _collapse(str(cell)).casefold())

    if not any(parts):
        return None

    digest = hashlib.sha1("\x1f".join(parts).encode("utf-8")).hexdigest()
    length = int(identity.get("fallback_length") or 20)
    prefix = str(identity.get("fallback_prefix") or "tb")
    return f"{prefix}-{digest[:max(8, length)]}"


# --------------------------------------------------------------------------
# Persistence
# --------------------------------------------------------------------------


def _chunked(items: Sequence[Any], size: int) -> Iterable[Sequence[Any]]:
    for start in range(0, len(items), size):
        yield items[start : start + size]


def _existing_rows(session: Any, request_ids: Sequence[str]) -> dict[str, tuple[Any, Any, Any]]:
    """Fetch {request_id: (service_type_raw, ingested_at, service_class)} for ids we are about to write."""
    from sqlalchemy import select

    existing: dict[str, tuple[Any, Any, Any]] = {}
    for chunk in _chunked(request_ids, _SELECT_CHUNK):
        statement = select(
            Request.request_id,
            Request.service_type_raw,
            Request.ingested_at,
            Request.service_class,
        ).where(Request.request_id.in_(list(chunk)))
        for request_id, service_type_raw, ingested_at, service_class in session.execute(statement):
            existing[request_id] = (service_type_raw, ingested_at, service_class)
    return existing


def _upsert(
    session: Any,
    records: Sequence[dict[str, Any]],
    rules: Mapping[str, Any],
) -> tuple[int, int, int]:
    """Upsert records on request_id. Returns (inserted, updated, service_type_conflicts).

    Reconciled in Python rather than in SQL so that two invariants hold on
    every re-ingest without a dialect-specific COALESCE dance:

    * ``ingested_at`` keeps its original first-seen value;
    * ``service_type_raw`` is immutable -- if a later export disagrees with what
      was stored, the stored value wins and the disagreement is counted. The
      derived ``service_class`` is then re-derived from the value that actually
      stays in the row, so the two can never drift apart.
    """
    if not records:
        return 0, 0, 0

    ids = [record["request_id"] for record in records]
    existing = _existing_rows(session, ids)

    inserted = updated = conflicts = 0
    for record in records:
        previous = existing.get(record["request_id"])
        if previous is None:
            inserted += 1
            continue
        updated += 1
        stored_raw, stored_ingested_at, _stored_class = previous
        if stored_ingested_at is not None:
            record["ingested_at"] = stored_ingested_at
        if stored_raw:
            if record["service_type_raw"] and record["service_type_raw"] != stored_raw:
                conflicts += 1
            record["service_type_raw"] = stored_raw
            record["service_class"] = classifier.classify_service(stored_raw, rules)

    dialect = session.get_bind().dialect.name
    insert_fn = None
    if dialect == "sqlite":
        from sqlalchemy.dialects.sqlite import insert as insert_fn  # type: ignore[no-redef]
    elif dialect == "postgresql":
        from sqlalchemy.dialects.postgresql import insert as insert_fn  # type: ignore[no-redef]

    if insert_fn is None:
        # Any other backend: correctness over speed.
        for record in records:
            session.merge(Request(**record))
        session.flush()
        return inserted, updated, conflicts

    for chunk in _chunked(records, _UPSERT_CHUNK):
        statement = insert_fn(Request).values([dict(item) for item in chunk])
        statement = statement.on_conflict_do_update(
            index_elements=[Request.request_id],
            set_={name: statement.excluded[name] for name in _UPSERT_COLUMNS if name != "request_id"},
        )
        session.execute(statement)
    session.flush()
    return inserted, updated, conflicts


def _write_run(run_id: str, **values: Any) -> None:
    """Create or update the runs row. Own session, so a failed ingest still logs."""
    try:
        with get_session() as session:
            run = session.get(Run, run_id)
            if run is None:
                run = Run(run_id=run_id)
                run.started_at = values.get("started_at") or utcnow()
                session.add(run)
            for name, value in values.items():
                if value is None:
                    continue
                if name in {"window_start", "window_end"} and getattr(run, name) is not None:
                    continue  # the acquirer's window is authoritative
                setattr(run, name, value)
    except Exception as exc:  # pragma: no cover - the run log must never mask the real error
        logger.error("could not record run %s: %s", run_id, exc)


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------


def ingest(
    xlsx_path: Path | str,
    run_id: str,
    account_id: str = DEFAULT_ACCOUNT_ID,
    company_id: str | None = None,
) -> IngestResult:
    """Ingest one archived Request Log export for one company. Idempotent.

    ``company_id`` stamps every row and the runs row, so a multi-company
    install can pull two tenants into one database and keep them apart
    afterwards. ``account_id`` is the old name for it and still works.

    Raises :class:`HeaderValidationError` when the export's headers no longer
    match config/schema.yaml (nothing is stored), and :class:`IngestionError`
    for any other failure. Both paths record the run as ``failed`` and emit a
    ``pipeline_failure`` event first -- silence is never treated as success.
    """
    global _schema_ready

    company_id = company_id if company_id is not None else account_id
    account_id = company_id

    path = Path(xlsx_path)
    started_at = utcnow()
    version = rules_version()
    result = IngestResult(run_id=run_id, xlsx_path=str(path), rules_version=version)

    if not _schema_ready:
        # Idempotent, and it turns "no such table: requests" into a working
        # first run for seed/manual invocations.
        init_db()
        _schema_ready = True

    _write_run(
        run_id,
        account_id=account_id,
        status="started",
        started_at=started_at,
        rules_version=version,
        source_file=str(path),
    )

    try:
        if not path.is_file():
            raise IngestionError(f"export file does not exist: {path}")

        schema = get_schema()
        rules = get_rules()
        parsed = _parse_sheet(path, schema, rules, run_id, company_id)

        result.rows_read = parsed.rows_read
        result.rows_rejected = parsed.rows_rejected
        result.reject_reasons = dict(parsed.reject_reasons)
        result.unmapped_statuses = dict(parsed.unmapped_statuses)
        result.unparsed_datetimes = parsed.unparsed_datetimes
        result.unparsed_amounts = parsed.unparsed_amounts
        result.duplicate_request_ids = parsed.duplicate_request_ids
        result.header_row = parsed.header_row
        result.matched_headers = dict(parsed.matched)
        result.unmapped_fields = list(parsed.unmapped_fields)

        default_class = classifier.default_service_class(rules)
        result.unclassified_count = sum(
            1 for record in parsed.records if record.get("service_class") == default_class
        )

        with get_session() as session:
            inserted, updated, conflicts = _upsert(session, parsed.records, rules)

        result.rows_inserted = inserted
        result.rows_updated = updated
        result.rows_upserted = inserted + updated
        result.service_type_conflicts = conflicts

        offered = [
            record["offered_at"] for record in parsed.records if record.get("offered_at") is not None
        ]

        _check_degraded(result, schema)

        _write_run(
            run_id,
            account_id=account_id,
            status=result.status,
            row_count=result.rows_upserted,
            rules_version=version,
            finished_at=utcnow(),
            source_file=str(path),
            window_start=min(offered) if offered else None,
            window_end=max(offered) if offered else None,
        )

        logger.info(
            "ingest %s: read %d, upserted %d (%d new, %d updated), rejected %d, "
            "unmapped statuses %d, unclassified %d",
            run_id,
            result.rows_read,
            result.rows_upserted,
            result.rows_inserted,
            result.rows_updated,
            result.rows_rejected,
            result.unmapped_status_count,
            result.unclassified_count,
        )
        return result

    except HeaderValidationError as exc:
        _write_run(
            run_id,
            account_id=account_id,
            status="failed",
            rules_version=version,
            finished_at=utcnow(),
            error_message=str(exc)[:4000],
            source_file=str(path),
        )
        emit_event(
            PIPELINE_FAILURE,
            {
                "stage": "ingestion",
                "reason": "header_mismatch",
                "severity": "high",
                "run_id": run_id,
                "account_id": account_id,
                "xlsx_path": str(path),
                "error": str(exc),
                "diff": exc.diff,
            },
        )
        raise

    except Exception as exc:
        message = f"{type(exc).__name__}: {exc}"
        _write_run(
            run_id,
            account_id=account_id,
            status="failed",
            rules_version=version,
            finished_at=utcnow(),
            error_message=message[:4000],
            source_file=str(path),
        )
        emit_event(
            PIPELINE_FAILURE,
            {
                "stage": "ingestion",
                "reason": "ingest_failed",
                "severity": "high",
                "run_id": run_id,
                "account_id": account_id,
                "xlsx_path": str(path),
                "error": message,
            },
        )
        logger.exception("ingest %s failed", run_id)
        if isinstance(exc, IngestionError):
            raise
        raise IngestionError(message) from exc


def _check_degraded(result: IngestResult, schema: Mapping[str, Any]) -> None:
    """Mark the run partial and alert when too much of the export was unusable.

    An ingest that stored every row but could not understand a tenth of the
    statuses produced numbers nobody should act on. It is not a crash, so it
    must not be reported as a clean success either.
    """
    try:
        threshold = float(schema.get("unknown_status_abort_threshold") or 0.10)
    except (TypeError, ValueError):
        threshold = 0.10

    if result.rows_read <= 0:
        return

    problems: list[str] = []
    unmapped_ratio = result.unmapped_status_count / result.rows_read
    rejected_ratio = result.rows_rejected / result.rows_read

    if unmapped_ratio > threshold:
        problems.append(
            f"{result.unmapped_status_count}/{result.rows_read} rows "
            f"({unmapped_ratio:.0%}) had a status config/schema.yaml does not map: "
            + ", ".join(
                f"{name!r} x{count}"
                for name, count in sorted(
                    result.unmapped_statuses.items(), key=lambda item: -item[1]
                )[:8]
            )
        )
    if rejected_ratio > threshold:
        problems.append(
            f"{result.rows_rejected}/{result.rows_read} rows ({rejected_ratio:.0%}) "
            "were rejected: "
            + ", ".join(f"{reason} x{count}" for reason, count in result.reject_reasons.items())
        )

    if not problems:
        return

    result.status = "partial"
    result.warnings.extend(problems)
    emit_event(
        PIPELINE_FAILURE,
        {
            "stage": "ingestion",
            "reason": "degraded_ingest",
            "severity": "high",
            "run_id": result.run_id,
            "xlsx_path": result.xlsx_path,
            "error": "ingest completed but the data is not trustworthy: " + "; ".join(problems),
            "detail": result.as_dict(),
        },
    )
    logger.error("ingest %s degraded: %s", result.run_id, "; ".join(problems))
